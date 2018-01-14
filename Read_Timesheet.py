import os
import openpyxl
import json
from openpyxl.utils import get_column_letter


class FindFiles:

    def __init__(self):
        self._patch = "C:\\Users\\bVd\\Desktop\\"
        self._user = "Bastiaan Van Denabeele"
        self._files = dict()


    def get_folder_list(self):
        #returns a list of all the files in the selected folder
        return os.listdir(self._patch)

    def find_excel(self):
        # returns and updates a dict with the als key the week and year of the file
        # as value the pacth of the located file
        for file in self.get_folder_list():
            if "Timesheet - {} - Week".format(self._user)in file:
                key = file.split(" - ")[2].split(".")[0]
                self._files.update({key: self._patch+file})
        return self._files


class ReadTimeSheets:

    def __init__(self):
        self._file_list_dict = FindFiles().find_excel()
        self._time_sheet_input = dict()

    def get_files_dir(self):
        return self._file_list_dict

    def get_sheets(self):
        # returns a dict with the als key the week and year of the file
        # as value the sheet 'Timesheet' of the excel file
        returndict = dict()
        for keys, value in self._file_list_dict.items():
            returndict.update({keys: openpyxl.load_workbook(value).get_sheet_by_name('Timesheet')})
        return returndict

    def get_sheet_input(self):
        # returns a dict with the als key the week and year of the file
        # as value a dict
        # In that dict the key are the coordinates of the cells en the value the values of the cells
        sheet_inputs = dict()
        for key, value in self.get_sheets().items():
            sheet_inputs.update({key: dict()})
            for rowOfCellObjects in value['A5':self.get_last_entry_timesheet(value)]:
                for cellObj in rowOfCellObjects:
                    if cellObj.value is not None:
                        sheet_inputs.get(key).update({cellObj.coordinate: cellObj.value})

        return sheet_inputs

    def get_last_entry_timesheet(self, sheet_object: openpyxl):
        # find the maximum range of data in the sheet
        return self.get_last_entry_column(sheet_object, row=4) + str(self.get_last_entry_row(sheet_object, start_row=5))


    @staticmethod
    def get_last_entry_row(sheet_object: openpyxl, start_row=1, column=1):
        # find the row number of the last entry in the given column
        row = start_row
        last_entry_row = row
        while sheet_object.cell(column=column, row=row).value is not None:
            last_entry_row = row
            row += 1
        return last_entry_row

    @staticmethod
    def get_last_entry_column(sheet_object: openpyxl, start_column=1, row=1):
        # find the column number of the last entry the given row
        column = start_column
        last_entry_column = column
        while sheet_object.cell(column=column, row=row).value is not None:
            last_entry_column = column
            column += 1
        return get_column_letter(last_entry_column)

    def get_type_of_activity_data(self, file_week=''):

        type_of_activity = dict()

        # get the path of the gives file or get a random path in the dict self._file_list_dict
        if file_week in self._file_list_dict.keys():
            patch = self._file_list_dict.get(file_week)
        else:
            patch = list(self._file_list_dict.values())[0]

        activity_sheet = openpyxl.load_workbook(patch).get_sheet_by_name('TypeOfActivity')

        for i in range(2,self.get_last_entry_row(activity_sheet)):
            column = 2

            while activity_sheet.cell(column=column, row=i).value is not None \
                    and activity_sheet.cell(column=column, row=i).value != 1:
                column += 1

            if activity_sheet.cell(column=column, row=i).value == 1:
                type_of_activity.update({activity_sheet.cell(column=1, row=i).value:
                                         activity_sheet.cell(column=column, row=1).value})
        return type_of_activity

    def update_type_of_activity_json(self):
        with open("type_of_activity.json", 'w') as outfile:
            json.dump(self.get_type_of_activity_data(), outfile)


if __name__ == "__main__":
    rt = ReadTimeSheets()
    rt.update_type_of_activity_json()
